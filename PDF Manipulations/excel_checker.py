from pathlib import Path
import openpyxl

root_path = "/mnt/c/Users/Yash/Desktop/Desktop/Aero Telcom Intern Work/Task 18(PDF manipulations)/PDF Test"
# print(root_path)
# print(os.getcwd())
contents = Path(root_path).glob("*")
# print(contents)
for content in contents:
    content = str(content)
    if(content.find(".")!=-1):
        continue
    else:               
        contents1 = Path(content).glob("*")
        for c1 in contents1:
            c1 = str(c1)  
            if(c1.find('~')!=-1):
                continue
            if(c1.endswith("_PRODUCTS.xlsx")):
                excel_name = c1

        wb_obj = openpyxl.load_workbook(excel_name)
        sheet_obj = wb_obj.active
        rows = sheet_obj.max_row
        # print(rows)
        cols = sheet_obj.max_column
        # print(cols)
        print(excel_name)
        for i in range(1,cols+1):
            col_name = sheet_obj.cell(1,i).value
            col_name = str(col_name)
            if(len(col_name)>28):
                print(f"Cell location (1,{i}) and value ({sheet_obj.cell(1,i).value}), exceeds 28 characters")
        wb_obj.close()
        print("\n")
        print(40*"=")
