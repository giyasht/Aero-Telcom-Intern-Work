import os
from pathlib import Path
import time
import csv
import pdftotext
import openpyxl
from wand.image import Image as wi
import PyPDF2
from PyPDF2 import PdfFileMerger

root_path = "/mnt/c/Users/Yash/Desktop/Desktop/Aero Telcom Intern Work/Task 18(PDF manipulations)/PDF Test"
mergingpdfpath = "/mnt/c/Users/Yash/Desktop/Desktop/Aero Telcom Intern Work/Task 18(PDF manipulations)/Contact Us.pdf"
# print(root_path)
# print(os.getcwd())
contents = Path(root_path).glob("*")
# print(contents)
for content in contents:
    content = str(content)
    if(content.find(".")!=-1):
        continue
    else:
        pdfs_folder = content + "/CsvsPDFDownloaded"
        img_dir = content + "/Images/Images"
        newpdf_dir = content + "/Updated PDFs"
        os.chdir(content)
        if not os.path.exists(newpdf_dir):
            os.mkdir(newpdf_dir)
        
        pdfs_content = Path(pdfs_folder).glob("*")
        
        for pdf in pdfs_content:
            pdf = str(pdf)
            if(pdf.endswith(".pdf")):
                pdfWriter = PyPDF2.PdfFileWriter()
                try:
                    pdfReader = PyPDF2.PdfFileReader(pdf, strict=False)

                    if pdfReader.getNumPages() > 1:
                        for i in range(pdfReader.getNumPages() - 1):
                            page = pdfReader.getPage(i)
                            pdfWriter.addPage(page)

                    newpdfname = pdf.split("/")[-1]
                    with open(newpdf_dir + '/' + newpdfname, "wb") as outpdf:
                        pdfWriter.write(outpdf)
                        outpdf.close()

                    merger = PdfFileMerger()
                    merger.append(newpdf_dir + "/" + newpdfname)
                    merger.append(mergingpdfpath)
                    os.remove(newpdf_dir + "/" + newpdfname)
                    merger.write(newpdf_dir + "/" + newpdfname.replace(".pdf", "") + ".pdf")
                    merger.close()

                except Exception as e:
                    print(pdf)
                    print(e)
                
        # pdfs = Path(pdfs_folder).glob("*")
        # for p in pdfs:
        #     print(p)
        contents1 = Path(content).glob("*")
        for c1 in contents1:
            c1 = str(c1)  
            if(c1.find('~')!=-1):
                continue
            if(c1.endswith("_PRODUCTS.xlsx")):
                excel_name = c1
            if(c1.endswith("ImageDetails.csv")):
                img_csv = c1
            if(c1.endswith("downloaded_pdf.csv")):
                pdf_csv = c1
        imgdetails={}
        pdfdetails={}
        with open(img_csv) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            lc=0
            for row in csv_reader:
                if(lc==0):
                    lc+=1
                    continue
                else:
                    if(row[0]):
                        imgdetails[row[0]]=row[1]
                lc+=1
            csv_file.close()
        with open(pdf_csv) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            lc=0
            for row in csv_reader:
                if(lc==0):
                    lc+=1
                    continue
                else:
                    if(row[0]):
                        pdfdetails[row[0]]=row[1]
                lc+=1
            csv_file.close()
        # print(prodimgnames)
        # print(prodpdfnames)
        # print(excel_name)
        images_csv = open(img_csv,'a',newline='')
        csv_writer = csv.writer(images_csv)

        wb_obj = openpyxl.load_workbook(excel_name)
        sheet_obj = wb_obj.active
        rows = sheet_obj.max_row
        # print(rows)
        cols = sheet_obj.max_column
        # print(cols)
        for i in range(2,rows+1):
            mfr_no = sheet_obj.cell(i,1).value
            # print(mfr_no)
            if(mfr_no in pdfdetails.keys()):
                if(mfr_no in imgdetails.keys()):
                    x=1
                else:
                    pdfpath = pdfs_folder+ "/"+ pdfdetails[mfr_no]+".pdf"
                    try:
                        PDFfile = wi(filename=pdfpath)
                        for img in PDFfile.sequence[:1]:
                            Image = wi(image=img, resolution=920)
                            iname = img_dir + '/' + pdfdetails[mfr_no].replace(".pdf", "")
                            Image.save(filename=iname + ".jpg")
                            Image.close()
                        print((iname.split("/"))[-1])
                        sheet_obj.cell(i,8).value = pdfdetails[mfr_no]
                        csv_writer.writerow([mfr_no,(iname.split("/"))[-1]])
                    except Exception as e:
                        print(pdfpath)
                        print(e)
                try:
                    with open(pdfs_folder+"/"+pdfdetails[mfr_no]+".pdf","rb") as f:
                        pdfdata = pdftotext.PDF(f)
                    text = pdfdata[0]
                    text = str(text)
                    text = text.strip()
                    # print(text)
                    sheet_obj.cell(i,cols+1).value = text
                except Exception as e:
                    print(pdfs_folder+"/"+pdfdetails[mfr_no]+".pdf")
                    print(e)
        wb_obj.save(excel_name)
        # sheet_obj.cell(row = 1, column = 26).value=34
        # wb_obj.save(excel_name)
        # print(sheet_obj.cell(row = 1, column = 26).value)
        images_csv.close()
        wb_obj.close()
